import pandas as pd
from geopy.geocoders import GoogleV3
from openpyxl import load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink

# 
# Google API Key #
GEOCODER = GoogleV3(api_key='#')

# https://stackoverflow.com/questions/50714216/how-to-correctly-access-geocode-api-response-in-python
# issue on address, not all cells contain correct address information. Using name instead
def get_map_url(hospital_name):
    location = GEOCODER.geocode(hospital_name, exactly_one=True)
    if location:
        return f"https://www.google.com/maps/?q={location.latitude},{location.longitude}"
    return "No URL Found"

# Row count
    """ def update_spreadsheet(input_file, output_file, max_rows=150):
        if not input_file or not output_file:
        print("Input or output file not specified. Exiting.")
        return """
# ignore. Kept breaking on cycle through sheets

def update_spreadsheet(input_file):
    # Open sheet (setting file names below)
    workbook = load_workbook(filename=input_file)
    sheet = workbook.active

    # https://stackoverflow.com/questions/23527887/getting-sheet-names-from-openpyxl
    #  make new sheet - no longer using this, update initial instead as loop was breaking
   # new_workbook = load_workbook(filename=input_file)
   # new_sheet = new_workbook.active

# https://stackoverflow.com/questions/29792134/how-we-can-use-iter-rows-in-python-openpyxl-package
# fixed issue of blank cell
    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=False):
        cell = row[0]
        if cell.value:
            # Make map holder `[Map_Me]` also setingt this in the sheet
            placeholder = '[Map_Me]'
            
            # Find the company address, have to split the cell value into lines
            lines = cell.value.split('\n')
            if len(lines) >= 1:
                hospital_name = lines[0]  # Hospital name from the first line
                
                # Get map URL
                new_map_url = get_map_url(hospital_name)
                
                # Prepare new cell values
                hospital_info = f"{lines[0]}\n{lines[1] if len(lines) > 1 else ''}"
                
                # Set hospital name and address in one cell
                sheet.cell(row=cell.row, column=1, value=hospital_info)
                
                # https://stackoverflow.com/questions/44811523/how-do-i-add-a-column-to-an-existing-excel-file-using-python
                # add map link in the next cell in next column
                hyperlink_cell = sheet.cell(row=cell.row, column=2)
                hyperlink_cell.value = 'Map_Me'
                hyperlink_cell.hyperlink = new_map_url
                hyperlink_cell.style = 'Hyperlink'
                
                print(f"Updated cell {cell.row} with clickable link.")
            else:
                print(f"No placeholder found in row {cell.row}")

    # Save the updated DataFrame to a new Excel file
    # changed, update existing file instead
    workbook.save(input_file)
    print(f"Spreadsheet saved as {input_file}")

                # Replace holder and add hyperlink
    """ if placeholder in cell.value:
                    # Create the hyperlink
                    new_sheet.cell(row=cell.row, column=1).value = cell.value.replace(placeholder, "")
                    new_sheet.cell(row=cell.row, column=1).hyperlink = new_map_url
                    new_sheet.cell(row=cell.row, column=1).style = 'Hyperlink'
                    print(f"Updated cell {cell.row} with clickable link.")
                else:
                    print(f"No placeholder found in row {cell.row}") """

                

    """ # Save the updated DataFrame to a new Excel file
    new_workbook.save(output_file)
    print(f"Spreadsheet saved as {output_file}") """

# File names
input_file = 'hospital_names_with_maps10.xlsx'
# output_file = 'hospital_names_with_maps10.xlsx'

# Update the spreadsheet
# removed output_file
update_spreadsheet(input_file)
